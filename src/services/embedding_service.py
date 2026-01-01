"""Embedding service for document processing."""
import hashlib
from typing import Optional, List
from uuid import UUID

from src.core.s3 import s3_client
from src.core.openai_client import openai_client
from src.core.pinecone_client import pinecone_client
from src.core.websocket import ws_manager
from src.core.database import db
from src.config import settings


class EmbeddingService:
    CHUNK_SIZE = 2000
    CHUNK_OVERLAP = 200

    @staticmethod
    async def extract_text(s3_key: str, file_type: str) -> Optional[str]:
        """Extract text from document."""
        content = await s3_client.get_file_content(s3_key)
        if not content:
            return None

        try:
            if file_type in ["txt", "md"]:
                return content.decode("utf-8", errors="ignore")

            elif file_type == "pdf":
                from PyPDF2 import PdfReader
                from io import BytesIO
                reader = PdfReader(BytesIO(content))
                return "\n".join(page.extract_text() or "" for page in reader.pages)

            elif file_type == "docx":
                from docx import Document
                from io import BytesIO
                doc = Document(BytesIO(content))
                return "\n".join(p.text for p in doc.paragraphs)

            elif file_type in ["xlsx", "xls"]:
                from openpyxl import load_workbook
                from io import BytesIO
                wb = load_workbook(BytesIO(content), read_only=True)
                texts = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        texts.append(" ".join(str(c) for c in row if c))
                return "\n".join(texts)

        except Exception as e:
            print(f"Text extraction error: {e}")

        return None

    @staticmethod
    def chunk_text(text: str) -> List[str]:
        """Split text into chunks."""
        chunks = []
        start = 0
        while start < len(text):
            end = start + EmbeddingService.CHUNK_SIZE
            chunks.append(text[start:end])
            start = end - EmbeddingService.CHUNK_OVERLAP
        return chunks

    @staticmethod
    async def process_document(document_id: UUID, org_id: Optional[UUID], s3_key: str,
                               file_type: str, user_id: str, upload_id: str):
        """Process document: extract text and generate embeddings."""
        try:
            # Update status
            db.admin.table("storage_nodes").update({
                "processing_status": "processing"
            }).eq("id", str(document_id)).execute()

            await ws_manager.send_upload_progress(user_id, upload_id, "extracting", 40)

            # Check if embeddings are enabled and file type is supported
            if not settings.ENABLE_EMBEDDINGS or file_type not in settings.SUPPORTED_EMBEDDING_TYPES:
                db.admin.table("storage_nodes").update({
                    "processing_status": "completed",
                    "embedding_status": "skipped"
                }).eq("id", str(document_id)).execute()
                await ws_manager.send_upload_progress(user_id, upload_id, "complete", 100, str(document_id))
                return

            # Extract text
            text = await EmbeddingService.extract_text(s3_key, file_type)
            if not text:
                db.admin.table("storage_nodes").update({
                    "processing_status": "completed",
                    "embedding_status": "failed"
                }).eq("id", str(document_id)).execute()
                await ws_manager.send_upload_progress(user_id, upload_id, "complete", 100, str(document_id))
                return

            await ws_manager.send_upload_progress(user_id, upload_id, "generating_embeddings", 60)

            # Chunk text
            chunks = EmbeddingService.chunk_text(text)

            # Generate embeddings
            embeddings = await openai_client.get_embeddings_batch(chunks)

            await ws_manager.send_upload_progress(user_id, upload_id, "storing_embeddings", 80)

            pinecone_namespace = str(org_id) if org_id else user_id # Consistent namespace
            # Prepare vectors for Pinecone
            vectors = []
            for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
                vectors.append({
                    'id': f"{document_id}_{i}",
                    'values': embedding,
                    'metadata': {
                        'document_id': str(document_id),
                        'org_id': pinecone_namespace, # Consistent identifier
                        'chunk_index': i,
                        'chunk_text': chunk[:500],
                    }
                })

            # Store in Pinecone
            await pinecone_client.upsert(vectors, pinecone_namespace)

            # Update document status
            db.admin.table("storage_nodes").update({
                "processing_status": "completed",
                "embedding_status": "completed"
            }).eq("id", str(document_id)).execute()

            await ws_manager.send_upload_progress(user_id, upload_id, "complete", 100, str(document_id))

        except Exception as e:
            print(f"Embedding error: {e}")
            db.admin.table("storage_nodes").update({
                "processing_status": "failed",
                "embedding_status": "failed"
            }).eq("id", str(document_id)).execute()
            await ws_manager.send_upload_progress(user_id, upload_id, "failed", 0, error=str(e))

    @staticmethod
    async def search(query: str, org_id: Optional[UUID], user_id: UUID, top_k: int = 10,
                     folder_id: Optional[UUID] = None) -> List[dict]:
        """Search documents using vector similarity."""
        # Get query embedding
        query_embedding = await openai_client.get_embedding(query)

        # Build filter
        filter_dict = None
        if folder_id:
            filter_dict = {'folder_id': str(folder_id)}

        pinecone_namespace = str(org_id) if org_id else str(user_id) # Consistent namespace
        # Query Pinecone
        results = await pinecone_client.query(query_embedding, pinecone_namespace, top_k, filter_dict)

        # Get document details
        doc_ids = list(set(r.metadata.get('document_id') for r in results if r.metadata))
        if not doc_ids:
            return []

        docs = db.admin.table("storage_nodes").select("*").in_("id", doc_ids).execute()
        doc_map = {d['id']: d for d in docs.data}

        return [
            {
                'score': r.score,
                'document': doc_map.get(r.metadata['document_id']),
                'chunk_text': r.metadata.get('chunk_text', '')
            }
            for r in results if r.metadata.get('document_id') in doc_map
        ]


embedding_service = EmbeddingService()
